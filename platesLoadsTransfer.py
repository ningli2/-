import openpyxl,datetime,csv,PIL

class Folder(object):
    
    def __init__(self,excelSPR,templateFile):
        self.excelFile=excelSPR
        self.templateFile=templateFile
    
    def getSPRexcel(self):
        wb=openpyxl.load_workbook(self.excelFile)
        sheet=wb.get_sheet_by_name('Sheet1')
        row_count=sheet.max_row
        column_count=sheet.max_column
        plateInformation=[]
        app={}
        for i in range(6,row_count):
            for j in range(2,column_count):
                app.setdefault(str(sheet.cell(row=5,column=j).value),
                                str(sheet.cell(row=i+1,column=j).value))
            plateInformation.append(app)
            app={}
        return plateInformation
        
    def getDiameter(self):
        wb=openpyxl.load_workbook(self.templateFile)
        sheet=wb.get_sheet_by_name('Sheet1')
        diameter=[]
        diameter.append(str(sheet.cell(row=1,column=18).value))
        diameter.append(str(sheet.cell(row=2,column=18).value))        
        return diameter    
    
    def fillTemplate(self):
        wb=openpyxl.load_workbook(self.templateFile)
        sheet=wb.get_sheet_by_name('Sheet1')   
        plateInformation=self.getSPRexcel() 
        plateInformationCopy=plateInformation.copy()
        for item in plateInformationCopy:
            if len(item['Vessel Diameter'])==0:
                plateInformation.remove(item)
        currentRow=10
        for j in range(len(self.getDiameter())):
            for i in range(len(plateInformation)):
                if plateInformation[i]['Vessel Diameter']==self.getDiameter()[j]:
                    name=plateInformation[i]['NAME'].split('-')
                    if len(name)==4:
                        supportName=name[3].split('(')
                        sheet.cell(row=currentRow,column=3).value=(
                        name[1]+'-'+name[2])
                        sheet.cell(row=currentRow,column=4).value=(
                        supportName[0])
                        sheet.cell(row=currentRow,column=5).value=int(float(
                        plateInformation[i]['Plate Level'])*1000)
                        sheet.cell(row=currentRow,column=6).value=int(float(
                        plateInformation[i]['Plate Width'])*1000)                
                        sheet.cell(row=currentRow,column=7).value=int(float(
                        plateInformation[i]['Plate Height'])*1000)    
                        sheet.cell(row=currentRow,column=8).value=int(float(
                        plateInformation[i]['Plate Thickness'])*1000)  
                        sheet.cell(row=currentRow,column=9).value=int(round(float(
                        plateInformation[i]['Plate Angle'])*180/3.14))  
                        currentRow +=1
        wb.save(r'C:\TEMP\platInformation.xlsx')
          

class Load(object):
    
    def __init__(self,loads,templateFile):
        self.loads=loads
        self.templateFile=templateFile
        
    def getLoads(self):
        loadsData=[]
        with open(self.loads) as loadsFile:
            loadsReader=csv.reader(loadsFile)
            loadsData=list(loadsReader)
        return loadsData
        
    def fillLoads(self):
        wb=openpyxl.load_workbook(self.templateFile)
        sheet=wb.get_sheet_by_name('Sheet1')   
        row_count=sheet.max_row
        loadsData=self.getLoads()
        print(loadsData)
        for j in range(10,row_count):
            for i in range(len(loadsData)):
                if loadsData[i][2]==str(sheet.cell(row=j,column=17).value): 
                    if loadsData[i][5][1]=='H':
                        for k in range(6):
                            if (isinstance(loadsData[i+13][3+k],str)
                                and loadsData[i+13][3+k].find(','))==-1:     
                                loadsData[i+13][3+k]=float(loadsData[i+13][3+k])*0.001
                            else:
                                loadsData[i+13][3+k]=float(str(loadsData[i+13][3+k]).replace(',','.'))
                        sheet.cell(row=j,column=10).value=loadsData[i+13][5]
                        sheet.cell(row=j,column=11).value=loadsData[i+13][4]
                        sheet.cell(row=j,column=12).value=loadsData[i+13][3] 
                        sheet.cell(row=j,column=13).value=loadsData[i+13][8]  
                        sheet.cell(row=j,column=14).value=loadsData[i+13][6] 
                        sheet.cell(row=j,column=15).value=loadsData[i+13][7]   
                        break
                    else:
                        for k in range(6):
                            if (isinstance(loadsData[i+13][3+k],str)
                                and loadsData[i+13][3+k].find(','))==-1:     
                                loadsData[i+13][3+k]=float(loadsData[i+13][3+k])*0.001
                            else:
                                loadsData[i+13][3+k]=float(str(loadsData[i+13][3+k]).replace(',','.'))
                        sheet.cell(row=j,column=10).value=loadsData[i+13][5]
                        sheet.cell(row=j,column=11).value=loadsData[i+13][3]
                        sheet.cell(row=j,column=12).value=loadsData[i+13][4] 
                        sheet.cell(row=j,column=13).value=loadsData[i+13][8]  
                        sheet.cell(row=j,column=14).value=loadsData[i+13][7] 
                        sheet.cell(row=j,column=15).value=loadsData[i+13][6]   
                        break
        png_loc = r'O:\Design_Piping\10@Script\coordinate.png'
        png_loc2 = r'O:\Design_Piping\10@Script\show.png'        
        my_png = openpyxl.drawing.image.Image(png_loc)
        my_png2 = openpyxl.drawing.image.Image(png_loc2)
        sheet.add_image(my_png,'J6')
        sheet.add_image(my_png2,'J30')                
        wb.save(r'C:\TEMP\updateWithLoads.xlsx')
#excelSPR=r'C:\TEMP\Plate.xlsx'
#templateFile=r'C:\TEMP\plateLoadsTemplate.xlsm'
#loads=r'C:\TEMP\loads.csv'
#excel=Folder(excelSPR,templateFile)
#print(excel.getSPRexcel())
#print(excel.getDiameter())
#excel.fillTemplate()
#loadFile=Load(loads,templateFile)
#print(loadFile.getLoads())
#print(loadFile.fillLoads())
